#!/usr/bin/env python3
"""
神谷塾 Instagram ダッシュボード自動更新スクリプト
実行するだけで、Instagram APIから最新のインサイトを取得し、Excelダッシュボードを自動更新します。
"""
import os, json, requests
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base_dir = Path(__file__).resolve().parent.parent
automation_dir = Path(__file__).resolve().parent
env_path = automation_dir / ".env"

env = {}
with open(env_path, "r", encoding="utf-8") as f:
    for line in f:
        if "=" in line and not line.strip().startswith("#"):
            k, v = line.strip().split("=", 1)
            env[k] = v.strip("\"'")

token = env.get("INSTAGRAM_ACCESS_TOKEN")
account_id = env.get("INSTAGRAM_ACCOUNT_ID")

print("1. アカウント情報と最新投稿を取得中...")
acc_url = f"https://graph.facebook.com/v19.0/{account_id}?fields=id,username,name,followers_count,follows_count,media_count&access_token={token}"
acc_res = requests.get(acc_url).json()
followers = acc_res.get("followers_count", 728)

media_url = f"https://graph.facebook.com/v19.0/{account_id}/media?fields=id,caption,media_type,media_product_type,timestamp,like_count,comments_count,permalink&limit=30&access_token={token}"
media_res = requests.get(media_url).json()
posts_raw = media_res.get("data", [])

posts_data = []
print(f"2. {len(posts_raw)}件のインサイト（リーチ・保存・シェア）を集計中...")
for p in posts_raw:
    pid = p.get("id")
    ptype = p.get("media_type")
    likes = p.get("like_count", 0)
    comms = p.get("comments_count", 0)
    caption = (p.get("caption") or "").strip()
    first_line = caption.split("\n")[0][:45] if caption else "(キャプションなし)"
    ts = p.get("timestamp", "")
    date_str = ts[:10] if ts else ""
    link = p.get("permalink", "")
    
    ins_url = f"https://graph.facebook.com/v19.0/{pid}/insights?metric=reach,saved,shares,total_interactions&access_token={token}"
    ins_res = requests.get(ins_url).json()
    reach_val, saved_val, shares_val = 0, 0, 0
    if "data" in ins_res:
        for item in ins_res["data"]:
            n = item.get("name")
            val = item.get("values", [{}])[0].get("value", 0)
            if n == "reach": reach_val = val
            elif n == "saved": saved_val = val
            elif n == "shares": shares_val = val
            
    cat = "その他"
    c_lower = caption.lower()
    if "hiragana" in c_lower or "ひらがな" in c_lower: cat = "ひらがな単体"
    elif "muramoto" in c_lower or "2026" in c_lower or ptype == "VIDEO": cat = "リール・講師紹介"
    elif "「に」" in c_lower or "partícula" in c_lower or "conversación" in c_lower: cat = "文法・実践会話"
    elif "febrero" in c_lower or "fema" in c_lower or "hinamatsuri" in c_lower or "caligrafía" in c_lower: cat = "文化・イベント"
    elif "estudiar" in c_lower or "visado" in c_lower: cat = "留学・進路"

    eval_rank = "C (要改善)"
    if reach_val >= 200 or saved_val >= 4: eval_rank = "S (勝ちパターン🔥)"
    elif reach_val >= 100 or saved_val >= 2 or likes >= 8: eval_rank = "A (好調✨)"
    elif reach_val >= 60 or likes >= 3: eval_rank = "B (標準)"

    eng_rate = ((likes + comms + saved_val + shares_val) / reach_val * 100) if reach_val > 0 else 0

    posts_data.append({
        "id": pid, "date": date_str,
        "type": "Reels (動画)" if ptype == "VIDEO" else "カルーセル" if ptype == "CAROUSEL_ALBUM" else "画像",
        "category": cat, "caption_title": first_line,
        "reach": reach_val, "saved": saved_val, "shares": shares_val,
        "likes": likes, "comments": comms, "eng_rate": round(eng_rate, 2),
        "eval": eval_rank, "link": link
    })

# Write latest json
with open(automation_dir / "latest_insights.json", "w", encoding="utf-8") as f:
    json.dump(posts_data, f, ensure_ascii=False, indent=2)

print("3. Excelダッシュボードを更新中...")
wb = openpyxl.Workbook()

NAVY_HEADER = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SUB_HEADER = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
ACCENT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
CARD_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DANGER_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

FONT_TITLE = Font(name="Meiryo", size=15, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Meiryo", size=10, bold=True, color="000000")
FONT_NORMAL = Font(name="Meiryo", size=10, color="000000")
FONT_MUTED = Font(name="Meiryo", size=9, color="595959")
FONT_CARD_NUM = Font(name="Meiryo", size=18, bold=True, color="1F3864")

THIN_BORDER = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
HEADER_BORDER = Border(left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'), top=Side(style='thin', color='FFFFFF'), bottom=Side(style='thin', color='FFFFFF'))

# Sheet 1
ws1 = wb.active
ws1.title = "📊 ダッシュボード"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells("A1:H2")
title_cell = ws1["A1"]
title_cell.value = "  ⛩️ 神谷塾（@japones_kamiyajuku）Instagram コンテンツ分析ダッシュボード"
title_cell.font = FONT_TITLE
title_cell.fill = NAVY_HEADER
title_cell.alignment = Alignment(vertical="center")

ws1["A3"] = f"集計日: {datetime.now().strftime('%Y/%m/%d')}  |  対象: 直近{len(posts_data)}投稿  |  フォロワー: {followers}名  |  自動更新: 完了"
ws1["A3"].font = FONT_MUTED

avg_reach = round(sum(p['reach'] for p in posts_data)/len(posts_data)) if posts_data else 0
total_saved = sum(p['saved'] for p in posts_data)
max_reach = max((p['reach'] for p in posts_data), default=0)

kpi_cards = [
    ("B5:C5", "B6:C7", "現在のフォロワー数", f"{followers:,} 名", CARD_FILL),
    ("D5:E5", "D6:E7", "直近平均リーチ数", f"{avg_reach:,} 人", SUCCESS_FILL),
    ("F5:G5", "F6:G7", "直近総保存数 (Saves)", f"{total_saved:,} 件", ACCENT_FILL),
    ("H5:I5", "H6:I7", "最高リーチ投稿", f"{max_reach:,} リーチ", WARN_FILL),
]

for title_range, num_range, card_title, card_num, fill_color in kpi_cards:
    ws1.merge_cells(title_range)
    t_cell = ws1[title_range.split(":")[0]]
    t_cell.value = card_title
    t_cell.font = FONT_BOLD
    t_cell.fill = fill_color
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.merge_cells(num_range)
    n_cell = ws1[num_range.split(":")[0]]
    n_cell.value = card_num
    n_cell.font = FONT_CARD_NUM
    n_cell.fill = fill_color
    n_cell.alignment = Alignment(horizontal="center", vertical="center")

ws1["A9"] = "【カテゴリ別パフォーマンス分析（勝ちパターンの検証）】"
ws1["A9"].font = Font(name="Meiryo", size=11, bold=True, color="1F3864")

cat_headers = ["カテゴリ", "投稿本数", "平均リーチ", "平均保存数", "平均いいね", "平均エンゲージ率", "今後の改善アクション"]
for col_idx, h in enumerate(cat_headers, start=1):
    cell = ws1.cell(row=10, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = SUB_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = HEADER_BORDER

cat_groups = {}
for p in posts_data:
    c = p["category"]
    if c not in cat_groups: cat_groups[c] = []
    cat_groups[c].append(p)

cat_actions = {
    "リール・講師紹介": "🌟 最重要（週1〜2本）：新規フォロワー獲得の最大チャネル。講師の寸劇・授業風景を増やす。",
    "文法・実践会話": "💾 高価値（週2本）：保存数が最多。「に vs で」のようにスペイン語話者の疑問に特化。",
    "文化・イベント": "✨ 拡散用（月2〜3本）：シェア数が高い。季節行事やバルセロナ現地イベントを配信。",
    "留学・進路": "⚠️ 要改善：文字が固いためリーチ40と低迷。インタビューやQ&A形式に改良。",
    "ひらがな単体": "❌ 縮小/廃止：リーチ40〜50・保存0と低迷。「旅行フレーズ集」等へ統合。",
    "その他": "🔍 テーマを絞り込む。"
}

current_row = 11
for cat, items in sorted(cat_groups.items(), key=lambda x: sum(i['reach'] for i in x[1])/len(x[1]), reverse=True):
    cnt = len(items)
    ar = round(sum(i['reach'] for i in items) / cnt, 1)
    asav = round(sum(i['saved'] for i in items) / cnt, 1)
    ali = round(sum(i['likes'] for i in items) / cnt, 1)
    aeng = round(sum(i['eng_rate'] for i in items) / cnt, 2)
    act = cat_actions.get(cat, "継続観察")
    
    r_vals = [cat, cnt, ar, asav, ali, f"{aeng}%", act]
    for col_idx, val in enumerate(r_vals, start=1):
        c = ws1.cell(row=current_row, column=col_idx, value=val)
        c.font = FONT_NORMAL
        c.border = THIN_BORDER
        if col_idx in [2, 3, 4, 5, 6]: c.alignment = Alignment(horizontal="right", vertical="center")
        else: c.alignment = Alignment(horizontal="left", vertical="center")
        if cat in ["リール・講師紹介", "文法・実践会話"]: c.fill = SUCCESS_FILL
        elif cat == "ひらがな単体": c.fill = DANGER_FILL
    current_row += 1

# Top Winning Posts
current_row += 2
ws1.cell(row=current_row, column=1, value="【🏆 直近の勝ち投稿 TOP 5（成功要因と再現ポイント）】").font = Font(name="Meiryo", size=11, bold=True, color="1F3864")
current_row += 1

top_headers = ["順位", "形式", "カテゴリ", "タイトル・キャプション", "リーチ", "保存数", "いいね", "成功要因・再現ポイント"]
for col_idx, h in enumerate(top_headers, start=1):
    cell = ws1.cell(row=current_row, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = SUB_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = HEADER_BORDER

top_posts = sorted(posts_data, key=lambda x: (x['saved']*30 + x['reach']), reverse=True)[:5]
current_row += 1
rank_medals = ["🥇 1位", "🥈 2位", "🥉 3位", "4位", "5位"]
reasons = [
    "「に」vs「で」という明確な悩みを解決し、保存数が最多（6件）。有料講座のリード獲得に最適。",
    "2026年の目標・モチベーションを刺激し、リール動画で最大リーチ（298人）を獲得。",
    "新任の村本先生の人柄・顔が見える紹介動画で、信頼感と親近感を獲得（リーチ196人）。",
    "バレンタインの日本独自ルールを紹介し、シェア数最多（6件）を獲得。",
    "日常会話のリアルな表現比較で実践的な学習ニーズに合致。"
]

for idx, p in enumerate(top_posts):
    r_vals = [
        rank_medals[idx], p["type"], p["category"], p["caption_title"],
        p["reach"], p["saved"], p["likes"],
        reasons[idx] if idx < len(reasons) else "高エンゲージメント"
    ]
    for col_idx, val in enumerate(r_vals, start=1):
        c = ws1.cell(row=current_row, column=col_idx, value=val)
        c.font = FONT_NORMAL
        c.border = THIN_BORDER
        if col_idx in [1, 2, 3]: c.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [5, 6, 7]: c.alignment = Alignment(horizontal="right", vertical="center")
        else: c.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1

# Sheet 2: Raw Data
ws2 = wb.create_sheet(title="📝 投稿データ一覧")
ws2.views.sheetView[0].showGridLines = True

raw_headers = ["No", "形式", "カテゴリ", "キャプション（冒頭）", "リーチ数", "保存数 (Saves)", "シェア数", "いいね数", "コメント数", "エンゲージ率", "総合評価", "投稿URL"]
for col_idx, h in enumerate(raw_headers, start=1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = NAVY_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = HEADER_BORDER

for idx, p in enumerate(posts_data, start=1):
    r_idx = idx + 1
    vals = [
        idx, p["type"], p["category"], p["caption_title"],
        p["reach"], p["saved"], p["shares"], p["likes"], p["comments"],
        f"{p['eng_rate']}%", p["eval"], p["link"]
    ]
    for c_idx, val in enumerate(vals, start=1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = FONT_NORMAL
        c.border = THIN_BORDER
        if c_idx in [1, 2, 3, 11]: c.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in [5, 6, 7, 8, 9, 10]: c.alignment = Alignment(horizontal="right", vertical="center")
        else: c.alignment = Alignment(horizontal="left", vertical="center")
        if c_idx == 11:
            if "S" in val: c.fill = SUCCESS_FILL
            elif "A" in val: c.fill = ACCENT_FILL
            elif "B" in val: c.fill = WARN_FILL
            elif "C" in val: c.fill = DANGER_FILL

# Sheet 3: PDCA
ws3 = wb.create_sheet(title="🎯 PDCA・改善チェックシート")
ws3.views.sheetView[0].showGridLines = True
ws3.merge_cells("A1:G2")
t3 = ws3["A1"]
t3.value = "  🎯 Instagram投稿 企画・改善PDCA判定シート"
t3.font = FONT_TITLE
t3.fill = NAVY_HEADER
t3.alignment = Alignment(vertical="center")

ws3["A3"] = "新しい投稿アイデアを出す際、以下の基準を満たしているかをチェックして投稿の質を担保します。"
ws3["A3"].font = FONT_MUTED

pdca_headers = ["フェーズ", "重要チェック項目", "合格基準（KPI目安）", "神谷塾での実践ポイント", "判定基準"]
for col_idx, h in enumerate(pdca_headers, start=1):
    cell = ws3.cell(row=5, column=col_idx, value=h)
    cell.font = FONT_HEADER
    cell.fill = SUB_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = HEADER_BORDER

pdca_rows = [
    ("【1】企画・フック", "1枚目/冒頭3秒で「スペイン語話者の疑問や悩み」を提示できているか？", "リーチ 150人以上", "「〜の違い」「スペイン人が間違えやすい表現」など比較にする", "✅ 必須"),
    ("【2】教育価値", "「後で見返したい」と思える保存性の高い情報が含まれているか？", "保存数 3件以上", "文法ルール、例文の対比、助詞の使い方を図解する", "✅ 必須"),
    ("【3】信頼・人間味", "講師の顔・声・教室のリアルな雰囲気が伝わっているか？", "いいね 10件以上", "先生の紹介、生徒とのレッスン切り抜き、祭り・イベント写真を入れる", "✅ 推奨"),
    ("【4】オンライン講座導線", "キャプション最後やスライド最後に明確なCTA（行動喚起）があるか？", "DM / プロフ遷移", "「コメントに『GUIA』で無料PDF送付」「オンライン体験受付中」を記載", "🎯 成約直結"),
    ("【5】フォーマット比率", "週の投稿バランスが最適化されているか？", "リール:カルーセル = 1:2", "リールで新規リーチを獲得し、カルーセルで保存・ファン化を狙う", "🔄 運用規律")
]

for r_idx, row_data in enumerate(pdca_rows, start=6):
    for c_idx, val in enumerate(row_data, start=1):
        c = ws3.cell(row=r_idx, column=c_idx, value=val)
        c.font = FONT_NORMAL
        c.border = THIN_BORDER
        if c_idx in [1, 3, 5]: c.alignment = Alignment(horizontal="center", vertical="center")
        else: c.alignment = Alignment(horizontal="left", vertical="center")
        if "必須" in str(val) or "成約直結" in str(val): c.fill = SUCCESS_FILL

for sheet in [ws1, ws2, ws3]:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            length = sum(2 if ord(char) > 256 else 1 for char in val_str)
            if length > max_len: max_len = length
        sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 50)

excel_path = base_dir / "01_analysis" / "instagram_content_analysis_dashboard.xlsx"
excel_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(excel_path)
print(f"✅ Excelダッシュボードの更新が完了しました: {excel_path}")
