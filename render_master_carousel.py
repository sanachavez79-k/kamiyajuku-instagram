import asyncio
import csv
import re
from pathlib import Path
from playwright.async_api import async_playwright
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "generated_assets"
ASSETS_DIR.mkdir(parents=True, exist_ok=True)
LOGO_PATH = BASE_DIR / "assets" / "brand_logo_main.png"

PHOTO_MONDAY = BASE_DIR / "assets" / "photo_monday.jpg"
PHOTO_WEDNESDAY = BASE_DIR / "assets" / "photo_wednesday.jpg"
PHOTO_FRIDAY = BASE_DIR / "assets" / "photo_friday.jpg"

MADRID_TZ = ZoneInfo("Europe/Madrid")

# 週替わりフォールバック用コンテンツプール
WEEKLY_CONTENT_POOL = {
    "LUNES": [
        {
            "pillar": "JLPT文法・重要助詞 (に vs で)",
            "tag_text": "JLPT N5 / N4 GRAMÁTICA ⛩️",
            "bg_primary": "#EEF4EA", "brand_deep": "#1B5E20", "accent_main": "#E8822A", "accent_light": "#DCEBD6",
            "student_photo": PHOTO_MONDAY, "photo_position": "center 20%",
            "title_html": '¿Dices "Tokyo <span style="color: #E8822A; text-decoration: underline;">NI</span>" o "Tokyo <span style="color: #1B5E20; text-decoration: underline;">DE</span>"? 🇯🇵❌',
            "subtitle": "El error con las partículas「に」y「で」que el 90% comete en el JLPT.",
            "hero_left": {"char": "に", "color": "#E8822A", "desc": "Estancia / Destino"},
            "hero_right": {"char": "で", "color": "#1B5E20", "desc": "Acción Activa"},
            "rule1": {"badge": "に (NI)", "badge_bg": "#E8822A", "title": "Lugar de Estancia / Destino", "desc": "Indica DÓNDE está una persona/objeto o a dónde vas.", "ja": '<ruby>東京<rt>とうきょう</rt></ruby>に <ruby>住<rt>す</rt></ruby>んでいます。', "es": "Vivo EN Tokio (Estancia fija)."},
            "rule2": {"badge": "で (DE)", "badge_bg": "#1B5E20", "title": "Lugar de Acción Activa", "desc": "Indica DÓNDE ocurre una actividad o evento dinámico.", "ja": 'レストランで <ruby>食<rt>た</rt></ruby>べます。', "es": "Como EN un restaurante (Acción)."},
            "q1_ja": '<ruby>図書館<rt>としょかん</rt></ruby>（ &nbsp;&nbsp;&nbsp;&nbsp; ）<ruby>本<rt>ほん</rt></ruby>を <ruby>読<rt>よ</rt></ruby>みます。', "q1_es": "Toshokan ( ) hon o yomimasu.",
            "q2_ja": '<ruby>机<rt>つくえ</rt></ruby>の <ruby>上<rt>うえ</rt></ruby>（ &nbsp;&nbsp;&nbsp;&nbsp; ）<ruby>猫<rt>ねこ</rt></ruby>が います。', "q2_es": "Tsukue no ue ( ) neko ga imasu.",
            "a1_text": "Respuesta: で (DE)", "a1_desc": '¡Porque "leer" es una acción activa en el lugar!',
            "a2_text": "Respuesta: に (NI)", "a2_desc": '¡Porque "estar/haber" indica existencia estática!',
            "cheat_t1": "¿Hay movimiento / acción activa?", "cheat_b1": '➔ 食べる, 勉強する, 買う = で (DE)',
            "cheat_t2": "¿Es estancia, estar o destino?", "cheat_b2": '➔ 住む, いる, ある, 行く = に (NI)',
            "dm_keyword": "JLPT", "dm_gift": "<b>Guía PDF gratuita de partículas</b> + <b>Test de Nivel</b>"
        },
        {
            "pillar": "JLPT文法・重要助詞 (は vs が)",
            "tag_text": "JLPT N5 / N4 GRAMÁTICA ⛩️",
            "bg_primary": "#EEF4EA", "brand_deep": "#1B5E20", "accent_main": "#E8822A", "accent_light": "#DCEBD6",
            "student_photo": PHOTO_MONDAY, "photo_position": "center 20%",
            "title_html": '¿Diferencia real entre <span style="color: #E8822A; text-decoration: underline;">WA (は)</span> y <span style="color: #1B5E20; text-decoration: underline;">GA (が)</span>? 🇯🇵🧠',
            "subtitle": "La duda más común explicada con la regla de 'Tema vs Sujeto & Clima'.",
            "hero_left": {"char": "は", "color": "#E8822A", "desc": "Tema Principal"},
            "hero_right": {"char": "が", "color": "#1B5E20", "desc": "Sujeto / Clima y Fenómenos"},
            "rule1": {"badge": "は (WA)", "badge_bg": "#E8822A", "title": "Presenta el Tema ('En cuanto a...')", "desc": "Pone el foco en lo que viene DESPUÉS de la partícula.", "ja": '<ruby>私<rt>わたし</rt></ruby>は カルロスです。', "es": "En cuanto a mí, soy Carlos."},
            "rule2": {"badge": "が (GA)", "badge_bg": "#1B5E20", "title": "Sujeto & Fenómenos del Clima / Naturaleza", "desc": "Pone el foco en el sujeto y se usa SIEMPRE para fenómenos del clima y naturaleza.", "ja": '<ruby>雨<rt>あめ</rt></ruby>が <ruby>降<rt>ふ</rt></ruby>る / <ruby>雷<rt>かみなり</rt></ruby>が <ruby>鳴<rt>な</rt></ruby>る', "es": "Llueve / Hay truenos. (Hechos y clima que percibes)"},
            "q1_ja": '<ruby>雨<rt>あめ</rt></ruby>（ &nbsp;&nbsp;&nbsp;&nbsp; ）<ruby>降<rt>ふ</rt></ruby>っています。', "q1_es": "Ame ( ) futte imasu. (Está lloviendo)",
            "q2_ja": '<ruby>今日<rt>きょう</rt></ruby>（ &nbsp;&nbsp;&nbsp;&nbsp; ）いい <ruby>天気<rt>てんき</rt></ruby>ですね。', "q2_es": "Kyou ( ) ii tenki desu ne. (Hoy hace buen tiempo)",
            "a1_text": "Respuesta: が (GA)", "a1_desc": "¡El clima y fenómenos naturales (lluvia, truenos, viento) llevan SIEMPRE が!",
            "a2_text": "Respuesta: は (WA)", "a2_desc": "¡Porque 'Hoy (今日)' es el tema sobre el que estamos hablando!",
            "cheat_t1": "¿Clima o preguntas con 誰 (quién), 何 (qué)?", "cheat_b1": '➔ 雨が降る, 雷が鳴る = が (GA)',
            "cheat_t2": "¿Contrastes o temas de conversación?", "cheat_b2": '➔ Siempre llevan は (WA)',
            "dm_keyword": "JLPT", "dm_gift": "<b>Guía PDF de Partículas N5/N4</b> + <b>Test de Nivel</b>"
        }
    ],
    "MIERCOLES": [
        {
            "pillar": "日常会話・リアル表現 (大丈夫の4つの意味)",
            "tag_text": "JAPONÉS REAL 🇯🇵",
            "bg_primary": "#FFFBEB", "brand_deep": "#B45309", "accent_main": "#D97706", "accent_light": "#FEF3C7",
            "student_photo": PHOTO_WEDNESDAY, "photo_position": "center 25%",
            "title_html": 'Los 4 significados de <span style="color: #D97706; text-decoration: underline;">だいじょうぶ (Daijoubu)</span> 🤯🇯🇵',
            "subtitle": "¡No solo significa 'Estoy bien'! Aprende a usarlo como un verdadero nativo.",
            "hero_single": {
                "kanji": "大丈夫？",
                "color": "#D97706",
                "romaji": "DAIJOUBU?",
                "desc": "¡La palabra más versátil y confusa del japonés cotidiano!"
            },
            "rule1": {"badge": "Significado 1 & 2", "badge_bg": "#D97706", "title": "OK (De acuerdo) / No gracias (Rechazo)", "desc": "Para aceptar con cortesía o para decir 'No gracias' en tiendas.", "ja": "これで 大丈夫です / 袋は 大丈夫です", "es": "Así está bien (OK) / Sin bolsa está bien (No gracias)."},
            "rule2": {"badge": "Significado 3 & 4", "badge_bg": "#B45309", "title": "¿Estás bien? / Sin problemas de salud", "desc": "Para preguntar por el estado de alguien o confirmar que estás bien.", "ja": "大丈夫ですか？ ➔ はい、大丈夫です！", "es": "¿Te encuentras bien? ➔ ¡Sí, todo bien!"},
            "q1_ja": "店員:「レジ袋は ご利用ですか？」 ➔ 客:「（ &nbsp;&nbsp;&nbsp;&nbsp; ）」", "q1_es": "Situación 1: En la caja de una tienda o supermercado.",
            "q1_opt_a": "A. 大丈夫です",
            "q1_opt_b": "B. ごめんなさい",
            "q2_ja": "友人が 転んだ時 ➔ 「（ &nbsp;&nbsp;&nbsp;&nbsp; ）！？」", "q2_es": "Situación 2: Cuando un amigo se tropieza en la calle.",
            "q2_opt_a": "A. 大丈夫！？",
            "q2_opt_b": "B. すみません！？",
            "a1_text": "Respuesta: A. 大丈夫です", "a1_desc": "¡La forma más natural y educada de decir 'No gracias' en tiendas!",
            "a2_text": "Respuesta: A. 大丈夫！？", "a2_desc": "¡La pregunta clave para saber si alguien necesita ayuda!",
            "cheat_t1": "¿Para rechazar con cortesía en tiendas?", "cheat_b1": "➔ 大丈夫です (No gracias / Estoy bien)",
            "cheat_t2": "¿Para confirmar que no hay problema?", "cheat_b2": "➔ 大丈夫です (Todo bien / OK)",
            "dm_keyword": "JLPT", "dm_gift": "<b>Guía Definitiva de Partículas JLPT (PDF)</b>"
        }
    ],
    "VIERNES": [
        {
            "pillar": "日本留学・ビザ (ビザ申請タイムライン)",
            "tag_text": "ESTUDIAR EN JAPÓN ✈️",
            "bg_primary": "#F0FDF4", "brand_deep": "#15803D", "accent_main": "#2E7D32", "accent_light": "#DCFCE7",
            "student_photo": PHOTO_FRIDAY, "photo_position": "center 20%",
            "title_html": '¿Quieres estudiar en Japón? <span style="color: #2E7D32; text-decoration: underline;">Calendario de Visa</span> 🇯🇵✈️',
            "subtitle": "Timeline exacto desde España para no perder las convocatorias oficiales.",
            "hero_left": {"char": "ビザ", "color": "#2E7D32", "desc": "Visado de Estudiante"},
            "hero_right": {"char": "準備", "color": "#15803D", "desc": "Paso a Paso"},
            "rule1": {"badge": "Paso 1 (5-6 meses antes)", "badge_bg": "#2E7D32", "title": "Elección de Escuela y CoE", "desc": "Seleccionar ciudad y preparar certificados bancarios de solvencia.", "ja": '<ruby>書類<rt>しょるい</rt></ruby>の <ruby>準備<rt>じゅんび</rt></ruby>を します。', "es": "Preparación de documentos oficiales y matrícula."},
            "rule2": {"badge": "Paso 2 (1-2 meses antes)", "badge_bg": "#15803D", "title": "Emisión del CoE y Visado", "desc": "Inmigración aprueba tu CoE y el Consulado emite tu visa.", "ja": 'ビザが <ruby>発給<rt>はっきゅう</rt></ruby>されます。', "es": "Emisión oficial del visado en el pasaporte."},
            "q1_ja": '留学生ビザの 申請は 出発の 何ヶ月前から？', "q1_es": "¿Con cuántos meses de antelación se tramita la visa de estudiante?",
            "q2_ja": '留学ビザで アルバイトは できる？', "q2_es": "¿Se puede trabajar con visa de estudiante en Japón?",
            "a1_text": "Respuesta: 5 a 6 meses de antelación", "a1_desc": "¡Porque Inmigración de Japón tarda hasta 3 meses en revisar el CoE!",
            "a2_text": "Respuesta: Sí, hasta 28 horas por semana", "a2_desc": "¡Con el permiso oficial de actividades que Kamiya Juku te ayuda a tramitar!",
            "cheat_t1": "📅 Convocatoria de Abril (Primavera)", "cheat_b1": "➔ Documentos en Octubre - Noviembre",
            "cheat_t2": "📅 Convocatoria de Octubre (Otoño)", "cheat_b2": "➔ Documentos en Abril - Mayo",
            "dm_keyword": "VISA", "dm_gift": "<b>Guía Completa para Estudiar en Japón desde España</b> + <b>Asesoría</b>"
        }
    ]
}

def load_config_from_sheet(day_key="LUNES"):
    """
    02_planning/content_ideas_sheet.xlsx の該当曜日タブからステータスが【READY】になっている最新行を読み込む。
    【PUBLISHED】は厳格に除外（二度と再配信しない）。
    """
    excel_candidates = [
        BASE_DIR / "02_planning" / "content_ideas_sheet.xlsx",
        BASE_DIR.parent / "02_planning" / "content_ideas_sheet.xlsx",
        Path("/Users/sanakamiya/Library/CloudStorage/GoogleDrive-kamiyajuku.japones@gmail.com/マイドライブ/インスタグラム/02_planning/content_ideas_sheet.xlsx")
    ]

    target_excel = None
    for p in excel_candidates:
        if p.exists():
            target_excel = p
            break

    tab_aliases = {
        "LUNES": ["月曜_JLPT文法", "Monday_JLPT", "LUNES", "月曜"],
        "MIERCOLES": ["水曜_日常会話", "Wednesday_Conversation", "MIERCOLES", "水曜"],
        "VIERNES": ["金曜_日本留学・ビザ", "Friday_Study_Visa", "VIERNES", "金曜"]
    }
    target_tab_names = tab_aliases.get(day_key, ["月曜_JLPT文法"])

    if target_excel:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(target_excel, data_only=True)
            ws = None
            for name in target_tab_names:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break

            if ws is not None:
                # 5行目から順に走査し、READY な行を抽出（PUBLISHED は除外）
                for r in range(5, ws.max_row + 1):
                    no = ws.cell(row=r, column=1).value
                    status = str(ws.cell(row=r, column=2).value or "").strip().upper()
                    theme = str(ws.cell(row=r, column=3).value or "").strip()
                    notes = str(ws.cell(row=r, column=4).value or "").strip()
                    sched_date = str(ws.cell(row=r, column=5).value or "").strip()

                    if status == "READY" and theme:
                        print(f"📖 Loaded READY theme from Excel ({ws.title} Row {r}): {theme}")
                        
                        # デフォルトスタイル
                        if day_key == "LUNES":
                            bg_primary, brand_deep, accent_main, accent_light, photo = "#EEF4EA", "#1B5E20", "#E8822A", "#DCEBD6", PHOTO_MONDAY
                            tag_text = "JLPT N5 / N4 GRAMÁTICA ⛩️"
                            dm_kw = "JLPT"
                            dm_gift = "<b>Guía PDF de Partículas N5/N4</b> + <b>Test de Nivel</b>"
                        elif day_key == "MIERCOLES":
                            bg_primary, brand_deep, accent_main, accent_light, photo = "#FFFBEB", "#B45309", "#D97706", "#FEF3C7", PHOTO_WEDNESDAY
                            tag_text = "JAPONÉS REAL 🇯🇵"
                            dm_kw = "JAPONES"
                            dm_gift = "<b>Guía de Expresiones Clave para Viajar a Japón</b> + <b>Audio</b>"
                        else:
                            bg_primary, brand_deep, accent_main, accent_light, photo = "#F0FDF4", "#15803D", "#2E7D32", "#DCFCE7", PHOTO_FRIDAY
                            tag_text = "ESTUDIAR EN JAPÓN ✈️"
                            dm_kw = "VISA"
                            dm_gift = "<b>Guía Completa para Estudiar en Japón desde España</b> + <b>Asesoría</b>"

                        # 既存プールとのマッチング
                        pool = WEEKLY_CONTENT_POOL.get(day_key, [])
                        for item in pool:
                            if any(k in item.get("pillar", "") for k in theme.split()) or any(k in theme for k in ["は", "が", "に", "で", "ため", "よう", "だいじょうぶ", "すみません", "ビザ", "生活費"]):
                                res = dict(item)
                                res["pillar"] = f"{theme}"
                                res["_row_id"] = no or r
                                res["_excel_path"] = str(target_excel)
                                res["_excel_row"] = r
                                res["_excel_sheet"] = ws.title
                                return res

                        # 新規テーマ自動構成
                        return {
                            "pillar": f"{theme}",
                            "tag_text": tag_text,
                            "bg_primary": bg_primary, "brand_deep": brand_deep, "accent_main": accent_main, "accent_light": accent_light,
                            "student_photo": photo, "photo_position": "center 20%",
                            "title_html": f'Aprende <span style="color: {accent_main}; text-decoration: underline;">{theme}</span> en japonés 🇯🇵🧠',
                            "subtitle": f"Consejos y reglas esenciales: {notes}" if notes else "Aprende la regla definitiva en 30 segundos con Kamiya Juku.",
                            "hero_left": {"char": "重要", "color": accent_main, "desc": "Regla Clave"},
                            "hero_right": {"char": "実践", "color": brand_deep, "desc": "Ejemplo Real"},
                            "rule1": {
                                "badge": "Punto 1", "badge_bg": accent_main,
                                "title": "Regla y Uso Principal", "desc": f"Explicación para {theme}",
                                "ja": f"{theme}の 使い方", "es": f"Uso correcto de {theme} en japonés natural."
                            },
                            "rule2": {
                                "badge": "Punto 2", "badge_bg": brand_deep,
                                "title": "Consejo y Caso Especial", "desc": notes if notes else "Ten cuidado con los errores más comunes.",
                                "ja": f"{notes}" if notes else "自然な 日本語の 表現", "es": "Expresión natural utilizada por nativos."
                            },
                            "q1_ja": f"¿Cómo se usa {theme}?", "q1_es": "Elige la opción más natural en una conversación.",
                            "q2_ja": "¿En qué situación es más adecuado?", "q2_es": "(a) Situación formal &nbsp;&nbsp; (b) Situación informal",
                            "a1_text": "Respuesta: ¡Opción correcta!", "a1_desc": f"¡Porque se adapta a la regla de {theme}!",
                            "a2_text": "Respuesta: (a) y (b)", "a2_desc": "¡Según el grado de cortesía y contexto!",
                            "cheat_t1": f"💡 Regla de oro para {theme}", "cheat_b1": f"➔ {notes}" if notes else "➔ Práctica activa y lectura diaria.",
                            "cheat_t2": "💼 ¿Quieres practicar con nativos?", "cheat_b2": "➔ Clases online en grupos reducidos de Kamiya Juku.",
                            "dm_keyword": dm_kw,
                            "dm_gift": dm_gift,
                            "_row_id": no or r,
                            "_excel_path": str(target_excel),
                            "_excel_row": r,
                            "_excel_sheet": ws.title
                        }
        except Exception as e:
            print(f"⚠️ Excel読み込みエラー: {e}")

    return None

def mark_post_as_published(day_key="LUNES", row_id=None, theme_name=None):
    """
    投稿完了後に 02_planning/content_ideas_sheet.xlsx の該当行の status を 'PUBLISHED' に更新
    """
    excel_candidates = [
        BASE_DIR / "02_planning" / "content_ideas_sheet.xlsx",
        BASE_DIR.parent / "02_planning" / "content_ideas_sheet.xlsx",
        Path("/Users/sanakamiya/Library/CloudStorage/GoogleDrive-kamiyajuku.japones@gmail.com/マイドライブ/インスタグラム/02_planning/content_ideas_sheet.xlsx")
    ]

    tab_aliases = {
        "LUNES": ["月曜_JLPT文法", "Monday_JLPT", "LUNES", "月曜"],
        "MIERCOLES": ["水曜_日常会話", "Wednesday_Conversation", "MIERCOLES", "水曜"],
        "VIERNES": ["金曜_日本留学・ビザ", "Friday_Study_Visa", "VIERNES", "金曜"]
    }
    target_tab_names = tab_aliases.get(day_key, ["月曜_JLPT文法"])

    for p in excel_candidates:
        if not p.exists():
            continue
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.load_workbook(p)
            ws = None
            for name in target_tab_names:
                if name in wb.sheetnames:
                    ws = wb[name]
                    break

            if ws is not None:
                for r in range(5, ws.max_row + 1):
                    no = ws.cell(row=r, column=1).value
                    theme = str(ws.cell(row=r, column=3).value or "").strip()
                    status = str(ws.cell(row=r, column=2).value or "").strip().upper()

                    match = False
                    if row_id and str(no) == str(row_id):
                        match = True
                    elif theme_name and (theme_name in theme or theme in theme_name) and status != "PUBLISHED":
                        match = True
                    elif not row_id and not theme_name and status == "READY":
                        match = True

                    if match:
                        cell = ws.cell(row=r, column=2)
                        cell.value = "PUBLISHED"
                        cell.font = Font(name="Helvetica Neue", size=11, color="6B7280")
                        cell.fill = PatternFill(fill_type=None)
                        print(f"✅ Marked as PUBLISHED in Excel ({p.name} -> {ws.title} Row {r})")
                        break

                wb.save(p)
        except Exception as e:
            print(f"⚠️ Failed to update published status in Excel {p}: {e}")

def get_current_week_config(day_key="LUNES", target_date=None):
    """
    1. content_ideas_sheet.csv から未投稿キュー（FIFO）を最優先で読み込み
    2. なければ週番号に基づいてフォールバックプールから選択
    """
    sheet_config = load_config_from_sheet(day_key)
    if sheet_config:
        return sheet_config

    pool = WEEKLY_CONTENT_POOL.get(day_key, WEEKLY_CONTENT_POOL["LUNES"])
    
    if target_date is None:
        now = datetime.now(MADRID_TZ)
        if now.weekday() in [6, 1, 3] and now.hour >= 20:
            target_date = now + timedelta(days=1)
        else:
            target_date = now

    current_week_num = target_date.isocalendar()[1]
    week_index = (current_week_num - 34) % len(pool)
    return pool[week_index]

# ==================== 漢字ふりがな（ルビ）辞書＆自動付与エンジン ====================
KANJI_RUBY_DICT = {
    '雨': 'あめ', '降': 'ふ', '雷': 'かみなり', '鳴': 'な',
    '私': 'わたし', '誰': 'だれ', '来': 'き', '今日': 'きょう', '天気': 'てんき',
    '東京': 'とうきょう', '京都': 'きょうと', '住': 'す', '食': 'た', '行': 'い', '勉強': 'べんきょう',
    '大丈夫': 'だいじょうぶ', '袋': 'ふくろ', '体調': 'たいちょう',
    '話': 'はな', '練習': 'れんしゅう', '風邪': 'かぜ', '買': 'か', '働': 'はたら', '貯金': 'ちょきん', '家': 'いえ',
    '日本': 'にほん', '日本語': 'にほんご', '学校': 'がっこう', '学生': 'がくせい', '留学生': 'りゅうがくせい',
    '準備': 'じゅんび', '書類': 'しょるい', '発給': 'はっきゅう', '申請': 'しんせい', '生活費': 'せいかつひ',
    '家賃': 'やちん', '部屋': 'へや', '時間': 'じかん', '毎日': 'まいにち', '朝': 'あさ', '夜': 'よる',
    '友達': 'ともだち', '友人': 'ゆうじん', '先生': 'せんせい', '使': 'つか', '方': 'かた', '意味': 'いみ',
    '表現': 'ひょうげん', '重要': 'じゅうよう', '自然': 'しぜん', '注意': 'ちゅうい', '会話': 'かいわ',
    '相づち': 'あいづち', '挨拶': 'あいさつ', '居酒屋': 'いざかや', '注文': 'ちゅうもん', '会計': 'かいけい',
    '店員': 'てんいん', '客': 'きゃく', '利用': 'りよう', '転': 'ころ', '時': 'とき',
    '開': 'あ', '閉': 'し', '入': 'はい', '出': 'で', '教': 'おし', '見': 'み', '聞': 'き'
}

def auto_add_furigana(text: str) -> str:
    """テキスト内の漢字に自動でふりがな（<ruby>タグ）を付与"""
    if not text:
        return ""

    # 既にrubyタグ化されている部分を保護しながら置換
    sorted_words = sorted(KANJI_RUBY_DICT.keys(), key=len, reverse=True)
    pattern = "(" + "|".join(re.escape(w) for w in sorted_words) + ")"

    # 簡易トークン分割（rubyタグ内は除外）
    parts = re.split(r"(<ruby>.*?</ruby>)", text)
    result = []
    for p in parts:
        if p.startswith("<ruby>"):
            result.append(p)
        else:
            def repl(m):
                w = m.group(1)
                return f"<ruby>{w}<rt>{KANJI_RUBY_DICT[w]}</rt></ruby>"
            result.append(re.sub(pattern, repl, p))

    return "".join(result)

def generate_master_day_html(day_key="LUNES", target_date=None):
    raw_c = get_current_week_config(day_key, target_date=target_date)
    c = dict(raw_c)
    if "rule1" in c and isinstance(c["rule1"], dict):
        c["rule1"] = dict(c["rule1"])
        c["rule1"]["ja"] = auto_add_furigana(c["rule1"].get("ja", ""))
    if "rule2" in c and isinstance(c["rule2"], dict):
        c["rule2"] = dict(c["rule2"])
        c["rule2"]["ja"] = auto_add_furigana(c["rule2"].get("ja", ""))
    if "q1_ja" in c:
        c["q1_ja"] = auto_add_furigana(c.get("q1_ja", ""))
    if "q2_ja" in c:
        c["q2_ja"] = auto_add_furigana(c.get("q2_ja", ""))
    if "q1_opt_a" in c:
        c["q1_opt_a"] = auto_add_furigana(c.get("q1_opt_a", ""))
    if "q1_opt_b" in c:
        c["q1_opt_b"] = auto_add_furigana(c.get("q1_opt_b", ""))
    if "q2_opt_a" in c:
        c["q2_opt_a"] = auto_add_furigana(c.get("q2_opt_a", ""))
    if "q2_opt_b" in c:
        c["q2_opt_b"] = auto_add_furigana(c.get("q2_opt_b", ""))
    if "a1_text" in c:
        c["a1_text"] = auto_add_furigana(c.get("a1_text", ""))
    if "a2_text" in c:
        c["a2_text"] = auto_add_furigana(c.get("a2_text", ""))
    if "cheat_b1" in c:
        c["cheat_b1"] = auto_add_furigana(c.get("cheat_b1", ""))
    if "cheat_b2" in c:
        c["cheat_b2"] = auto_add_furigana(c.get("cheat_b2", ""))

    if "hero_single" in c and c["hero_single"]:
        hs = dict(c["hero_single"])
        hs["kanji"] = auto_add_furigana(hs.get("kanji", ""))
        hero_html = f"""<div class="hero-single-card">
      <div class="hero-single-kanji" style="color: {hs.get('color', 'var(--brand-deep)')};">{hs['kanji']}</div>
      <div class="hero-single-romaji">{hs.get('romaji', '')}</div>
      <div class="hero-single-desc">{hs.get('desc', '')}</div>
    </div>"""
    else:
        hl = c.get('hero_left', {'color': 'var(--brand-deep)', 'char': '', 'desc': ''})
        hr = c.get('hero_right', {'color': 'var(--accent-main)', 'char': '', 'desc': ''})
        hero_html = f"""<div class="hero-box">
      <div class="hero-card">
        <div class="hero-kanji" style="color: {hl['color']};">{hl['char']}</div>
        <div class="hero-desc">{hl['desc']}</div>
      </div>
      <div class="hero-card">
        <div class="hero-kanji" style="color: {hr['color']};">{hr['char']}</div>
        <div class="hero-desc">{hr['desc']}</div>
      </div>
    </div>"""
    
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@500;700;800;900&family=Noto+Sans+JP:wght@500;700;900&display=swap');

  :root {{
    --bg-primary: {c['bg_primary']};
    --bg-card: #FFFFFF;
    --brand-deep: {c['brand_deep']};
    --accent-main: {c['accent_main']};
    --accent-light: {c['accent_light']};
    --text-main: #1A202C;
    --text-muted: #4A5568;
    --card-shadow: 0 16px 36px rgba(0, 0, 0, 0.06);
    --border-radius: 32px;
  }}

  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background-color: #333; font-family: 'Montserrat', 'Noto Sans JP', sans-serif; }}

  /* ふりがな（ルビ）スタイル */
  ruby {{
    ruby-align: center;
    ruby-position: over;
  }}
  rt {{
    font-size: 0.52em;
    color: var(--brand-deep);
    font-weight: 700;
    line-height: 1;
    transform: translateY(-2px);
  }}

  .slide {{
    width: 1080px; height: 1350px;
    background-color: var(--bg-primary);
    position: relative;
    padding: 75px 70px;
    display: flex; flex-direction: column; justify-content: space-between;
    overflow: hidden;
  }}

  /* ヘッダー */
  .slide-header {{
    display: flex; justify-content: space-between; align-items: center;
    border-bottom: 2px solid rgba(0,0,0,0.08); padding-bottom: 22px;
  }}
  .logo-badge {{
    display: flex; align-items: center; gap: 14px;
    font-weight: 900; font-size: 24px; color: var(--brand-deep);
  }}
  .logo-badge img {{
    height: 48px; width: auto; object-fit: contain;
  }}
  .category-pill {{
    background: var(--bg-card); color: var(--brand-deep);
    padding: 8px 22px; border-radius: 50px; font-size: 17px; font-weight: 800;
    box-shadow: 0 4px 12px rgba(0,0,0,0.04); letter-spacing: 1px;
  }}

  /* フッター */
  .slide-footer {{
    display: flex; justify-content: space-between; align-items: center;
    border-top: 2px solid rgba(0,0,0,0.08); padding-top: 20px;
    font-size: 20px; font-weight: 800; color: var(--text-muted);
  }}
  .swipe-indicator {{
    display: flex; align-items: center; gap: 10px; color: var(--brand-deep);
    animation: bounce 1.5s infinite;
  }}
  @keyframes bounce {{
    0%, 100% {{ transform: translateX(0); }}
    50% {{ transform: translateX(8px); }}
  }}

  /* 共通カード */
  .content-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 44px; box-shadow: var(--card-shadow); border: 2px solid rgba(0,0,0,0.03);
  }}

  /* スライド1: 表紙 */
  .cover-title {{
    font-size: 54px; font-weight: 900; line-height: 1.25; color: var(--text-main);
    margin-bottom: 20px;
  }}
  .cover-subtitle {{
    font-size: 25px; font-weight: 600; color: var(--text-muted); line-height: 1.45;
  }}
  .hero-box {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin: 35px 0;
  }}
  .hero-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 30px; text-align: center; box-shadow: var(--card-shadow);
    border: 2px solid rgba(0,0,0,0.03);
  }}
  .hero-kanji {{
    font-size: 72px; font-weight: 900; line-height: 1; margin-bottom: 8px;
  }}
  .hero-desc {{
    font-size: 20px; font-weight: 700; color: var(--text-muted);
  }}

  /* シングルヒーローカード（中央ドーン配置） */
  .hero-single-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 40px 30px; text-align: center; box-shadow: var(--card-shadow);
    border: 3px solid var(--accent-light); margin: 35px 0;
  }}
  .hero-single-kanji {{
    font-size: 88px; font-weight: 900; line-height: 1.15; margin-bottom: 8px;
  }}
  .hero-single-romaji {{
    font-size: 24px; font-weight: 800; color: var(--brand-deep); letter-spacing: 2px; margin-bottom: 10px;
  }}
  .hero-single-desc {{
    font-size: 22px; font-weight: 700; color: var(--text-muted); line-height: 1.4;
  }}

  /* スライド2: ルール解説 */
  .rule-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 38px 44px; margin-bottom: 28px; box-shadow: var(--card-shadow);
  }}
  .rule-badge {{
    display: inline-block; padding: 6px 18px; border-radius: 12px;
    color: #FFF; font-weight: 900; font-size: 20px; margin-bottom: 14px;
  }}
  .rule-title {{
    font-size: 30px; font-weight: 900; color: var(--text-main); margin-bottom: 10px;
  }}
  .rule-desc {{
    font-size: 22px; color: var(--text-muted); line-height: 1.4; margin-bottom: 18px;
  }}
  .example-box {{
    background: var(--bg-primary); border-radius: 18px; padding: 20px 24px;
    border-left: 6px solid var(--brand-deep);
  }}
  .example-ja {{
    font-size: 30px; font-weight: 900; color: var(--text-main); margin-bottom: 6px;
  }}
  .example-es {{
    font-size: 20px; font-weight: 600; color: var(--text-muted);
  }}

  /* スライド3: クイズ */
  .quiz-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 36px 42px; margin-bottom: 24px; box-shadow: var(--card-shadow);
  }}
  .quiz-q {{
    font-size: 30px; font-weight: 900; color: var(--text-main); margin-bottom: 8px; line-height: 1.35;
  }}
  .quiz-sub {{
    font-size: 21px; font-weight: 600; color: var(--text-muted); margin-bottom: 16px;
  }}
  .quiz-options {{
    display: flex; gap: 16px;
  }}
  .quiz-opt {{
    flex: 1; background: var(--bg-primary); padding: 18px; border-radius: 16px;
    text-align: center; font-size: 26px; font-weight: 900; color: var(--brand-deep);
    border: 2px dashed rgba(0,0,0,0.15);
  }}

  /* スライド4: クイズ正解 */
  .ans-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 36px 42px; margin-bottom: 24px; box-shadow: var(--card-shadow);
    border-left: 8px solid var(--accent-main);
  }}
  .ans-header {{
    display: flex; align-items: center; gap: 12px; margin-bottom: 10px;
  }}
  .ans-badge {{
    background: var(--accent-main); color: #FFF; padding: 4px 14px;
    border-radius: 8px; font-size: 18px; font-weight: 900;
  }}
  .ans-text {{
    font-size: 30px; font-weight: 900; color: var(--text-main);
  }}
  .ans-desc {{
    font-size: 22px; color: var(--text-muted); line-height: 1.45; margin-top: 10px;
  }}

  /* スライド5: チートシート */
  .cheat-container {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 44px; box-shadow: var(--card-shadow); margin-bottom: 20px;
  }}
  .cheat-row {{
    margin-bottom: 30px; padding-bottom: 25px; border-bottom: 2px dashed rgba(0,0,0,0.08);
  }}
  .cheat-row:last-child {{ margin-bottom: 0; padding-bottom: 0; border-bottom: none; }}
  .cheat-q {{
    font-size: 26px; font-weight: 800; color: var(--text-muted); margin-bottom: 10px;
  }}
  .cheat-a {{
    font-size: 36px; font-weight: 900; color: var(--text-main); line-height: 1.3;
  }}

  /* スライド6: 最終CTA & 生徒写真 */
  .cta-container {{
    display: flex; flex-direction: column; height: 100%; justify-content: space-between;
  }}
  .photo-frame {{
    width: 100%; height: 520px; border-radius: var(--border-radius);
    overflow: hidden; box-shadow: var(--card-shadow); position: relative;
  }}
  .photo-frame img {{
    width: 100%; height: 100%; object-fit: cover; object-position: {c['photo_position']};
  }}
  .photo-overlay {{
    position: absolute; bottom: 0; left: 0; right: 0;
    background: linear-gradient(transparent, rgba(0,0,0,0.7));
    padding: 30px; color: #FFF; font-size: 24px; font-weight: 800;
  }}
  .cta-card {{
    background: var(--bg-card); border-radius: var(--border-radius);
    padding: 40px; box-shadow: var(--card-shadow); text-align: center;
    border: 3px solid var(--accent-main);
  }}
  .cta-title {{
    font-size: 34px; font-weight: 900; color: var(--text-main); margin-bottom: 12px;
  }}
  .cta-sub {{
    font-size: 22px; font-weight: 600; color: var(--text-muted); margin-bottom: 24px; line-height: 1.4;
  }}
  .cta-badge {{
    display: inline-block; background: var(--accent-main); color: #FFF;
    font-size: 30px; font-weight: 900; padding: 14px 38px; border-radius: 50px;
    letter-spacing: 1px; box-shadow: 0 10px 24px rgba(232, 130, 42, 0.35);
  }}
</style>
</head>
<body>

<!-- SLIDE 1: Cover -->
<div class="slide" id="slide-1">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 KAMIYA JUKU</span>
    </div>
    <div class="category-pill">{c['tag_text']}</div>
  </div>

  <div>
    <h1 class="cover-title">{c['title_html']}</h1>
    <p class="cover-subtitle">{c['subtitle']}</p>
    {hero_html}
  </div>

  <div class="slide-footer">
    <span>Desliza para ver la regla 👉</span>
    <div class="swipe-indicator">1 / 6</div>
  </div>
</div>

<!-- SLIDE 2: Reglas -->
<div class="slide" id="slide-2">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 REGLAS CLAVE</span>
    </div>
    <div class="category-pill">{c['tag_text']}</div>
  </div>

  <div>
    <div class="rule-card">
      <div class="rule-badge" style="background: {c['rule1']['badge_bg']};">{c['rule1']['badge']}</div>
      <div class="rule-title">{c['rule1']['title']}</div>
      <div class="rule-desc">{c['rule1']['desc']}</div>
      <div class="example-box">
        <div class="example-ja">{c['rule1']['ja']}</div>
        <div class="example-es">{c['rule1']['es']}</div>
      </div>
    </div>

    <div class="rule-card">
      <div class="rule-badge" style="background: {c['rule2']['badge_bg']};">{c['rule2']['badge']}</div>
      <div class="rule-title">{c['rule2']['title']}</div>
      <div class="rule-desc">{c['rule2']['desc']}</div>
      <div class="example-box" style="border-left-color: {c['brand_deep']};">
        <div class="example-ja">{c['rule2']['ja']}</div>
        <div class="example-es">{c['rule2']['es']}</div>
      </div>
    </div>
  </div>

  <div class="slide-footer">
    <span>¿Lo has entendido? ¡Ponte a prueba! 👉</span>
    <div class="swipe-indicator">2 / 6</div>
  </div>
</div>

<!-- SLIDE 3: Quiz -->
<div class="slide" id="slide-3">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 MINI QUIZ</span>
    </div>
    <div class="category-pill">TEST RÁPIDO ✍️</div>
  </div>

  <div>
    <div class="quiz-card">
      <div class="quiz-q">Q1. {c['q1_ja']}</div>
      <div class="quiz-sub">{c['q1_es']}</div>
      <div class="quiz-options">
        <div class="quiz-opt">{c.get('q1_opt_a', 'A')}</div>
        <div class="quiz-opt">{c.get('q1_opt_b', 'B')}</div>
      </div>
    </div>

    <div class="quiz-card">
      <div class="quiz-q">Q2. {c['q2_ja']}</div>
      <div class="quiz-sub">{c['q2_es']}</div>
      <div class="quiz-options">
        <div class="quiz-opt">{c.get('q2_opt_a', 'A')}</div>
        <div class="quiz-opt">{c.get('q2_opt_b', 'B')}</div>
      </div>
    </div>
  </div>

  <div class="slide-footer">
    <span>Comprueba tus respuestas 👉</span>
    <div class="swipe-indicator">3 / 6</div>
  </div>
</div>

<!-- SLIDE 4: Respuestas -->
<div class="slide" id="slide-4">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 RESPUESTAS</span>
    </div>
    <div class="category-pill">SOLUCIÓN 🎯</div>
  </div>

  <div>
    <div class="ans-card" style="border-left-color: {c['accent_main']};">
      <div class="ans-header">
        <span class="ans-badge" style="background: {c['accent_main']};">Q1</span>
        <span class="ans-text">{c['a1_text']}</span>
      </div>
      <div class="ans-desc">{c['a1_desc']}</div>
    </div>

    <div class="ans-card" style="border-left-color: {c['brand_deep']};">
      <div class="ans-header">
        <span class="ans-badge" style="background: {c['brand_deep']};">Q2</span>
        <span class="ans-text">{c['a2_text']}</span>
      </div>
      <div class="ans-desc">{c['a2_desc']}</div>
    </div>
  </div>

  <div class="slide-footer">
    <span>Guarda el resumen en la siguiente 👉</span>
    <div class="swipe-indicator">4 / 6</div>
  </div>
</div>

<!-- SLIDE 5: Cheat Sheet -->
<div class="slide" id="slide-5">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 RESUMEN RÁPIDO</span>
    </div>
    <div class="category-pill">CHEAT SHEET 📌</div>
  </div>

  <div class="cheat-container">
    <div class="cheat-row">
      <div class="cheat-q">{c['cheat_t1']}</div>
      <div class="cheat-a">{c['cheat_b1']}</div>
    </div>
    <div class="cheat-row">
      <div class="cheat-q">{c['cheat_t2']}</div>
      <div class="cheat-a">{c['cheat_b2']}</div>
    </div>
  </div>

  <div class="slide-footer">
    <span>🎁 Regalo exclusivo en la última 👉</span>
    <div class="swipe-indicator">5 / 6</div>
  </div>
</div>

<!-- SLIDE 6: CTA & Foto Real -->
<div class="slide" id="slide-6">
  <div class="slide-header">
    <div class="logo-badge">
      <img src="file://{LOGO_PATH.resolve()}" alt="Logo">
      <span>神谷塾 COMUNIDAD</span>
    </div>
    <div class="category-pill">REGALO GRATIS 🎁</div>
  </div>

  <div class="cta-container" style="margin: 25px 0;">
    <div class="photo-frame">
      <img src="file://{c['student_photo'].resolve()}" alt="Estudiantes de Kamiya Juku">
      <div class="photo-overlay">
        ⛩️ Academia de Japonés en Barcelona & Online
      </div>
    </div>

    <div class="cta-card">
      <div class="cta-title">¿Quieres dominar el japonés real?</div>
      <div class="cta-sub">
        Envía un DM con la palabra <strong>"{c['dm_keyword']}"</strong> para recibir nuestra {c['dm_gift']} gratis.
      </div>
      <div class="cta-badge">📩 DM "{c['dm_keyword']}"</div>
    </div>
  </div>

  <div class="slide-footer">
    <span>⛩️ @japones_kamiyajuku</span>
    <div class="swipe-indicator">6 / 6</div>
  </div>
</div>

</body>
</html>
"""
    return html

async def render_day_carousel(day_key="LUNES", target_date=None):
    html_content = generate_master_day_html(day_key, target_date=target_date)
    temp_html_path = ASSETS_DIR / f"temp_master_{day_key}.html"
    
    with open(temp_html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    output_paths = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width": 1080, "height": 1350}, device_scale_factor=2)
        
        await page.goto(f"file://{temp_html_path.resolve()}", wait_until="networkidle")
        await page.wait_for_timeout(1000)

        for i in range(1, 7):
            slide_elem = page.locator(f"#slide-{i}")
            out_file = ASSETS_DIR / f"master_slide_{day_key}_{i}.jpg"
            await slide_elem.screenshot(path=str(out_file), type="jpeg", quality=95)
            output_paths.append(str(out_file))

        await browser.close()

    return output_paths

if __name__ == "__main__":
    for d in ["LUNES", "MIERCOLES", "VIERNES"]:
        cfg = get_current_week_config(d)
        print(f"=== {d} Config ===")
        print("  Pillar:", cfg.get("pillar"))
        print("  Row ID:", cfg.get("_row_id"))
        print("  CSV:", cfg.get("_csv_path"))
